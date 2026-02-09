#!/usr/bin/env python3
import argparse
import csv
import math
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.legend import Legend
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install it with: pip install openpyxl"
    ) from exc


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def die(message: str) -> None:
    print(f"error: {message}", file=sys.stderr)
    raise SystemExit(1)


def parse_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col].width = min(max_len + 2, 60)


def build_details_sheet(wb: Workbook, rows: List[Dict[str, str]], headers: List[str]) -> None:
    ws = wb.create_sheet("Details")
    ws.append(headers)

    for col_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"

    numeric_cols = [
        h
        for h in headers
        if h.startswith("exec_")
        or h.startswith("read_")
        or h.startswith("hit_")
        or h in {"exec_ms", "shared_read", "shared_hit", "shared_dirtied", "party_count", "service_count"}
    ]
    integer_cols = {"party_count", "service_count"}
    for col_index, header in enumerate(headers, start=1):
        if header in numeric_cols:
            for row_index in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_index, column=col_index)
                if cell.value == "":
                    continue
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    continue
            for row_index in range(2, ws.max_row + 1):
                fmt = "0" if header in integer_cols else "0.00"
                ws.cell(row=row_index, column=col_index).number_format = fmt

    if ws.max_row > 1:
        for header in numeric_cols:
            if header in integer_cols:
                continue
            col_idx = headers.index(header) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

    autosize_columns(ws)


def percentile(values: List[float], pct: float) -> float:
    if not values:
        return math.nan
    ordered = sorted(values)
    if pct <= 0:
        return ordered[0]
    if pct >= 100:
        return ordered[-1]
    rank = (pct / 100) * (len(ordered) - 1)
    low = int(math.floor(rank))
    high = int(math.ceil(rank))
    if low == high:
        return ordered[low]
    weight = rank - low
    return ordered[low] * (1 - weight) + ordered[high] * weight


def summarize(values: List[float]) -> Dict[str, float]:
    if not values:
        return {"avg": math.nan, "p50": math.nan, "p95": math.nan, "p99": math.nan}
    return {
        "avg": sum(values) / len(values),
        "p50": percentile(values, 50),
        "p95": percentile(values, 95),
        "p99": percentile(values, 99),
    }


def build_summary_sheet(wb: Workbook, rows: List[Dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "Summary"

    headers = ["variant", "count", "exec_avg", "exec_min", "exec_max", "exec_p50", "exec_p95", "exec_p99"]
    ws.append(headers)
    for col_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    grouped: Dict[str, Dict[str, List[float]]] = defaultdict(lambda: {"exec_avg": [], "exec_min": [], "exec_max": [], "exec_p50": [], "exec_p95": [], "exec_p99": []})
    for row in rows:
        variant = row.get("variant") or ""
        if not variant:
            continue
        for key in ["exec_avg", "exec_min", "exec_max", "exec_p50", "exec_p95", "exec_p99"]:
            value = row.get(key)
            try:
                parsed = float(value) if value not in (None, "", "None") else None
            except ValueError:
                parsed = None
            if parsed is not None:
                grouped[variant][key].append(parsed)

    for variant in sorted(grouped.keys()):
        bucket = grouped[variant]
        stats = {
            "exec_avg": summarize(bucket["exec_avg"])["avg"],
            "exec_min": summarize(bucket["exec_min"])["avg"],
            "exec_max": summarize(bucket["exec_max"])["avg"],
            "exec_p50": summarize(bucket["exec_p50"])["avg"],
            "exec_p95": summarize(bucket["exec_p95"])["avg"],
            "exec_p99": summarize(bucket["exec_p99"])["avg"],
        }
        count = max(len(bucket["exec_avg"]), len(bucket["exec_p50"]), len(bucket["exec_p95"]), len(bucket["exec_p99"]))
        ws.append(
            [
                variant,
                count,
                stats["exec_avg"],
                stats["exec_min"],
                stats["exec_max"],
                stats["exec_p50"],
                stats["exec_p95"],
                stats["exec_p99"],
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"

    for col_index in range(3, len(headers) + 1):
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.number_format = "0.00"

    autosize_columns(ws)

    if ws.max_row > 1:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Exec p50/p95/p99 by Variant"
        chart.y_axis.title = "variant"
        chart.x_axis.title = "ms"
        chart.legend = Legend()
        data = Reference(ws, min_col=6, max_col=8, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 18
        ws.add_chart(chart, "K2")

    if ws.max_row > 1:
        for col_idx in range(3, 9):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate Excel summary from summary.csv")
    parser.add_argument("summary_csv", help="Path to summary.csv")
    parser.add_argument("--out", help="Output .xlsx path (default: alongside input)")
    args = parser.parse_args()

    summary_path = Path(args.summary_csv)
    if not summary_path.is_file():
        die(f"summary.csv not found: {summary_path}")

    out_path = Path(args.out) if args.out else summary_path.with_suffix(".xlsx")

    rows = parse_csv(summary_path)
    if not rows:
        die("summary.csv has no rows")

    wb = Workbook()

    build_summary_sheet(wb, rows)
    details_headers = list(rows[0].keys())
    build_details_sheet(wb, rows, details_headers)

    wb.save(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
